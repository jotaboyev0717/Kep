import openpyxl

def create_workbook():
    workbook = openpyxl.Workbook()
    workbook.active.title = 'sheet1'
    workbook.create_sheet('sheet2')
    workbook.create_sheet('sheet3')
    return workbook