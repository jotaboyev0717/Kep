import openpyxl

def create_workbook():
    workbook = openpyxl.Workbook()
    workbook.create_sheet()
    workbook.create_sheet()
    return workbook